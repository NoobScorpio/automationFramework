import openpyxl


class XMLParser:
    def __init__(self, file, sheet):
        self.file = file
        self.xmlBook = openpyxl.load_workbook(file)
        self.xml = self.xmlBook[sheet]

    def read_data(self, row, col):
        return self.xml.cell(row=row, column=col).value

    def write_date(self, row, col, data):
        self.xml.cell(row=row, column=col).value = data
        self.xmlBook.save(self.file)

    def get_row_col(self) -> tuple[int, int]:
        return self.xml.max_row, self.xml.max_column
